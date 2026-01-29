from tkinter import *
from tkinter import messagebox
from openpyxl import load_workbook

screen = Tk()
screen.geometry('700x200')
screen.resizable(width=False, height=False)
screen.title('FixIt')
screen['bg'] = 'cyan'


def check():
    P = path0.get()
    C = column0.get()
    if P and C:
        messagebox.showinfo('Success', 'Вы успешно отправили данные! Закройте приложение.')
        global path
        path = P
        global column_letter
        column_letter = C
    if not P and C:
        messagebox.showerror('Error 0x1', 'Введите путь к файлу!')
    elif not C and P:
        messagebox.showerror('Error 0x2', 'Введите букву столбца!')
    if not P and not C:
        messagebox.showerror('Error 1x1', 'Введите данные!')


text_path = Label(text='Введите полный путь к файлу с названием и расширением', font='Comfortaa 15',
                  fg='#3d3d42',
                  bg='cyan')
path0 = Entry(screen,
             fg='#3d3d42',
             bg='white',
             relief='solid',
             width=50)

text_column = Label(text='Введите букву столбца, который нужно исправить', font='Comfortaa 15',
                    fg='#3d3d42',
                    bg='cyan')
column0 = Entry(screen,
                fg='#3d3d42',
                bg='white',
                relief='solid')
enter_column = Button(text='Отправить',
                      relief='solid',
                      bg='white',
                      command=check)

text_path.pack(padx=10, pady=10)
path0.pack()
text_column.pack(padx=10, pady=10)
column0.pack()
enter_column.pack(padx=10, pady=10)

screen.mainloop()

print(path)
path = path.strip()
column_letter = column_letter.strip()
file_to_fix = load_workbook(path)
sheet_file_to_fix = file_to_fix.active

correct_file = load_workbook(r"C:\Users\Zara\Desktop\курсовая\Правильные.xlsx", read_only=True)
sh_correct = correct_file.active
sp_corr = []
for i in range(2, 6):
    sp_corr.append(sh_correct["B" + str(i)].value)


def lev(A, B):
    F = [[((i + j) if i * j == 0 else 0) for j in range(len(B) + 1)] for i in range(len(A) + 1)]
    for i in range(1, len(A) + 1):
        for j in range(1, len(B) + 1):
            if A[i - 1] == B[j - 1]:
                F[i][j] = F[i - 1][j - 1]
            else:
                F[i][j] = 1 + min(F[i - 1][j], F[i][j - 1], F[i - 1][j - 1])
    return F[len(A)][len(B)]


row_count = sheet_file_to_fix.max_row

for i in range(2, row_count+1):
    a = sheet_file_to_fix[column_letter + str(i)].value
    for j in sp_corr:
        my_try = lev(a, j)
        if my_try <= 3:
            sheet_file_to_fix["B" + str(i)].value = j

            file_to_fix.save(path)
            #file_to_fix.save(r"C:\Users\Zara\Desktop\курсовая\Проба.xlsx")
    print(a)